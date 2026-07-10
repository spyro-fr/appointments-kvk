from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.config import SLOTS_PER_DAY, NUM_DAYS
from src.models import Player
from src.time_slots import generate_slot_labels


DAY_NAMES = ["Monday", "Tuesday", "Thursday"]
# columns per table: day 0 has 6 cols (Slot, Pseudo, Trigram, ID, Speedups, TrueGold)
# days 1 and 2 have 5 cols (Slot, Pseudo, Trigram, ID, Speedups)
DAY_COL_COUNTS = [6, 5, 5]


def ensure_template(template_path: Path) -> None:
    """Create a template with three side-by-side tables.

    Layout (columns): each table has DAY_COL_COUNTS[day] columns, separated by 1 empty column.

    Row 1: day name header (merged across the N columns of each table)
    Row 2: column headers
    Row 3+: slot labels and entries
    """
    if template_path.exists():
        return

    template_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule"

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    # compute base column for each day dynamically
    base = 1
    for day_index in range(NUM_DAYS):
        cols = DAY_COL_COUNTS[day_index]
        # merge cells for day name across cols
        ws.merge_cells(start_row=1, start_column=base, end_row=1, end_column=base + cols - 1)
        day_cell = ws.cell(row=1, column=base, value=DAY_NAMES[day_index] if day_index < len(DAY_NAMES) else f"Day {day_index + 1}")
        day_cell.font = header_font
        day_cell.fill = header_fill
        day_cell.alignment = Alignment(horizontal="center", vertical="center")

        # column headers in row 2
        if day_index == 0:
            headers = ["Slot", "Pseudo", "Trigram", "ID", "Number of speedup days", "Number of TrueGold"]
        else:
            headers = ["Slot", "Pseudo", "Trigram", "ID", "Number of speedup days"]

        for i, h in enumerate(headers):
            cell = ws.cell(row=2, column=base + i, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # advance base by cols + one separator
        base += cols + 1

    # create labels: first the previous-day 23:45, then the usual SLOTS_PER_DAY labels (English)
    labels = ["23:45 (previous day)"] + generate_slot_labels()
    for i, label in enumerate(labels, start=3):
        # write the slot label in each table's Slot column for visual alignment
        base = 1
        for day_index in range(NUM_DAYS):
            ws.cell(row=i, column=base, value=label)
            base += DAY_COL_COUNTS[day_index] + 1

    # set sensible column widths
    base = 1
    for day_index in range(NUM_DAYS):
        cols = DAY_COL_COUNTS[day_index]
        # slot
        ws.column_dimensions[ws.cell(row=2, column=base).column_letter].width = 16
        # pseudo
        ws.column_dimensions[ws.cell(row=2, column=base + 1).column_letter].width = 25
        # trigram
        ws.column_dimensions[ws.cell(row=2, column=base + 2).column_letter].width = 12
        # id
        ws.column_dimensions[ws.cell(row=2, column=base + 3).column_letter].width = 18
        if cols >= 5:
            ws.column_dimensions[ws.cell(row=2, column=base + 4).column_letter].width = 16
        if cols >= 6:
            ws.column_dimensions[ws.cell(row=2, column=base + 5).column_letter].width = 16
        base += cols + 1

    wb.save(template_path)


def write_schedule_per_days(
    schedules: list[dict[int, Player]],
    template_path: Path,
    output_path: Path,
) -> None:
    """
    Write three tables side-by-side on a single worksheet.

    Each table has columns as specified in DAY_COL_COUNTS.
    Tables are separated by one empty column.
    """
    ensure_template(template_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(template_path)
    ws = wb.worksheets[0]
    ws.title = "Schedule"

    # ensure only one sheet remains
    while len(wb.worksheets) > 1:
        wb.remove(wb.worksheets[-1])

    labels = ["23:45 (veille)"] + generate_slot_labels()
    # rewrite labels (defensive)
    base = 1
    for day_index in range(NUM_DAYS):
        for row_offset, label in enumerate(labels, start=3):
            ws.cell(row=row_offset, column=base, value=label)
        base += DAY_COL_COUNTS[day_index] + 1

    # for each day write Pseudo / Trigram / ID / speedups / (truegold)
    base = 1
    for day_index in range(NUM_DAYS):
        schedule = schedules[day_index]
        cols = DAY_COL_COUNTS[day_index]
        for row in range(len(labels)):
            row_num = row + 3
            player = schedule.get(row)
            if player:
                ws.cell(row=row_num, column=base + 1, value=player.pseudo)
                ws.cell(row=row_num, column=base + 2, value=player.alliance_trigram)
                ws.cell(row=row_num, column=base + 3, value=player.player_id)
                if cols >= 5:
                    ws.cell(row=row_num, column=base + 4, value=player.speedups)
                if cols >= 6:
                    ws.cell(row=row_num, column=base + 5, value=player.truegold)
            else:
                # ensure blanks if no player
                ws.cell(row=row_num, column=base + 1, value="")
                ws.cell(row=row_num, column=base + 2, value="")
                ws.cell(row=row_num, column=base + 3, value="")
                if cols >= 5:
                    ws.cell(row=row_num, column=base + 4, value="")
                if cols >= 6:
                    ws.cell(row=row_num, column=base + 5, value="")
        base += cols + 1

    wb.save(output_path)
